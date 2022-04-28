from concurrent.futures import ThreadPoolExecutor
import os
from openpyxl import load_workbook
from urllib.parse import urlparse

import consts
from img import TextAdder, TextParams

ARTICLE_FIELD_NAME = "Файл с расшеринием"
TEXT_FIELD_NAME = "Текст"
UPDOWN_FIELD_NAME = "Верх/Низ"
HOR_ALIGN_FIELD_NAME = "Центровка текста по горизонтали"
VERT_MARGIN_FIELD_NAME = "Сдвиг по вертикали в пикселях"
FONT_NAME_FIELD_NAME = "Шрифт"
BOLD_FIELD_NAME = "Жирность"
ADD_BLOCK_FIELD_NAME = "Доп блоком или нет"
BKGRD_FIELD_NAME = "Цвет фона"
FONT_SIZE_FIELD_NAME = "Размер шрифта"
COLOR_FIELD_NAME = "Цвет текста"

COLORS_SCHEME = {
    "черный": "black",
    "белый": "white",
    "серый": "gray",
    "красный": "red",
    "коричневый": "brown",
    "бордовый": "darkred",
    "оранжевый": "orange",
    "желтый": "yellow",
    "зеленый": "green",
    "голубой": "skyblue",
    "синий": "blue",
    "фиолетовый": "purple",
}

ALIGN_SCHEME = {-1: "left", 0: "center", 1: "right"}

XLSX_HEADERS = {
    ARTICLE_FIELD_NAME,
    TEXT_FIELD_NAME,
    UPDOWN_FIELD_NAME,
    HOR_ALIGN_FIELD_NAME,
    VERT_MARGIN_FIELD_NAME,
    FONT_NAME_FIELD_NAME,
    BOLD_FIELD_NAME,
    ADD_BLOCK_FIELD_NAME,
    BKGRD_FIELD_NAME,
    FONT_SIZE_FIELD_NAME,
    COLOR_FIELD_NAME,
}


class Adder:
    def __init__(self, file_name: str = None):
        self.file_name = file_name
        self._items = []
        self._path_name = os.path.split(file_name)[0]
        self._work_sheet = None
        self._file_headers: dict = dict()
        self._item_number: int = 1
        self._max_items_number: int = 0

    def open_file(self):
        work_book = load_workbook(self.file_name, read_only=True, data_only=True)
        self._work_sheet = work_book.active
        self._file_headers = self._get_headers()
        if not self._file_headers or not XLSX_HEADERS.issubset(self._file_headers):
            print(consts.FILE_OPEN_EXCEPTION_MSG)
            raise ValueError(consts.FILE_OPEN_EXCEPTION_MSG)
        self._max_items_number = self._work_sheet.max_row - 1
        self._read_data()

    def _get_headers(self) -> dict:
        headers = {}
        header_row = next(self._work_sheet.iter_rows(min_row=1, max_col=30))
        for header_number, header in enumerate(header_row):
            headers[header.value] = header_number
        return headers

    def _create_item(self, row):
        print(f"Чтение из Файла {self._item_number}/{self._max_items_number}")
        self._item_number += 1
        img_file = str(os.path.basename(urlparse(row[self._file_headers[ARTICLE_FIELD_NAME]].value).path))
        img_fn = os.path.join(self._path_name, img_file)
        result_path = os.path.join(self._path_name, "results")
        text = row[self._file_headers[TEXT_FIELD_NAME]].value
        params = TextParams(
            font_name=row[self._file_headers[FONT_NAME_FIELD_NAME]].value,
            font_size=row[self._file_headers[FONT_SIZE_FIELD_NAME]].value,
            bkrg=COLORS_SCHEME[row[self._file_headers[BKGRD_FIELD_NAME]].value],
            updown=row[self._file_headers[UPDOWN_FIELD_NAME]].value,
            hor_align=ALIGN_SCHEME[row[self._file_headers[HOR_ALIGN_FIELD_NAME]].value],
            vert_margin=row[self._file_headers[VERT_MARGIN_FIELD_NAME]].value,
            bold=bool(row[self._file_headers[BOLD_FIELD_NAME]].value),
            add_block=bool(row[self._file_headers[ADD_BLOCK_FIELD_NAME]].value),
            color=COLORS_SCHEME[row[self._file_headers[COLOR_FIELD_NAME]].value],
        )

        try:
            self._items.append(
                TextAdder(
                    file_name=img_fn,
                    result_path=result_path,
                    text=text,
                    params=params,
                )
            )
        except FileNotFoundError:
            print(f"Не удалось обрабоать строку {self._item_number}. Файл {img_fn} не найден")

    def _read_data(self):
        with ThreadPoolExecutor(max_workers=8) as executor:
            executor.map(self._create_item, self._work_sheet.iter_rows(min_row=2, max_col=30))

        self._item_number = 0

    def _change_image(self, image):
        print(f"[Обработано {self._item_number}/ {len(self._items)}] Добавление текста к {image.file_name}")
        try:
            image.add_text()
            self._item_number += 1
        except Exception as e:
            print(f"Ошибка обработки файла с ошибкой {e}")

    def process(self):
        with ThreadPoolExecutor(max_workers=8) as executor:
            executor.map(self._change_image, self._items)
        input("Нажмите что-нибудь чтобы завершить")


if __name__ == "__main__":
    import sys

    path_name, _ = os.path.split(sys.argv[0])
    filename = os.path.join(path_name, "images.xlsx")

    frozen = "not"
    if getattr(sys, "frozen", False):
        # we are running in a bundle
        frozen = "ever so"
        bundle_dir = sys._MEIPASS
    else:
        # we are running in a normal Python environment
        bundle_dir = os.path.dirname(os.path.abspath(__file__))

    adder = Adder(file_name=filename)
    adder.open_file()
    adder.process()
